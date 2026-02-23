from __future__ import annotations
"""
academic_modeling.py — ML + econometric modeling on batch-extracted feature datasets.

Models available:
  - OLS regression
  - Fixed effects panel model
  - Difference-in-differences
  - Random Forest / XGBoost
  - LDA topic modeling
  - Time series trend analysis
"""

import warnings
warnings.filterwarnings("ignore")

from typing import Dict, List, Any, Optional, Tuple
import pandas as pd
import numpy as np


# ════════════════════════════════════════════════════════════════════════════
# FEATURE EXTRACTION FROM BUNDLE
# ════════════════════════════════════════════════════════════════════════════

def bundle_to_feature_row(bundle: Dict, metadata: Dict) -> Dict[str, Any]:
    """
    Flatten a visual bundle into a single feature row for batch analysis.
    Returns a dict with ~50 features covering pacing, acoustics, NLP, structure.
    """
    row = {
        # Metadata
        "title":           metadata.get("title", ""),
        "type":            metadata.get("type", ""),
        "genre":           metadata.get("genre", ""),
        "year":            metadata.get("year", ""),
        "source":          metadata.get("source", ""),
        "url":             metadata.get("url", ""),
        "uploader":        metadata.get("uploader", ""),
    }

    sm = bundle.get("summary", {})
    scene_df = bundle.get("scene_df", pd.DataFrame())
    acoustic_df = bundle.get("acoustic_df", pd.DataFrame())
    transcript = bundle.get("transcript_text", "")

    # ── Scene structure ──────────────────────────────────────────────────────
    row["n_scenes"]            = sm.get("scenes", 0)
    row["duration_seconds"]    = sm.get("duration_seconds", 0)
    row["duration_minutes"]    = round(sm.get("duration_seconds", 0) / 60, 2)
    row["total_words"]         = sm.get("dialogue_words", 0)
    row["words_per_minute"]    = round(sm.get("dialogue_words", 0) / max(sm.get("duration_seconds", 1) / 60, 0.01), 2)
    row["words_per_scene"]     = round(sm.get("dialogue_words", 0) / max(sm.get("scenes", 1), 1), 2)

    if not scene_df.empty and "dialogue_words" in scene_df.columns:
        dw = scene_df["dialogue_words"]
        row["scene_words_mean"]   = round(float(dw.mean()), 2)
        row["scene_words_median"] = round(float(dw.median()), 2)
        row["scene_words_std"]    = round(float(dw.std()), 2)
        row["scene_words_max"]    = int(dw.max())
        row["scene_words_min"]    = int(dw.min())
        row["scene_words_cv"]     = round(float(dw.std() / dw.mean()) if dw.mean() > 0 else 0, 4)

        # Act-level pacing
        n = len(scene_df)
        act1 = scene_df.iloc[:int(n*0.25)]["dialogue_words"]
        act2 = scene_df.iloc[int(n*0.25):int(n*0.75)]["dialogue_words"]
        act3 = scene_df.iloc[int(n*0.75):]["dialogue_words"]
        row["act1_words_mean"] = round(float(act1.mean()), 2) if not act1.empty else 0
        row["act2_words_mean"] = round(float(act2.mean()), 2) if not act2.empty else 0
        row["act3_words_mean"] = round(float(act3.mean()), 2) if not act3.empty else 0
        row["pacing_acceleration"] = round(
            (row["act3_words_mean"] - row["act1_words_mean"]) / max(row["act1_words_mean"], 1), 4
        )

    else:
        for k in ["scene_words_mean","scene_words_median","scene_words_std","scene_words_max",
                  "scene_words_min","scene_words_cv","act1_words_mean","act2_words_mean",
                  "act3_words_mean","pacing_acceleration"]:
            row[k] = None

    # ── Acoustic features ────────────────────────────────────────────────────
    if not acoustic_df.empty:
        if "rms_energy" in acoustic_df.columns:
            e = acoustic_df["rms_energy"]
            row["energy_mean"]   = round(float(e.mean()), 6)
            row["energy_std"]    = round(float(e.std()), 6)
            row["energy_max"]    = round(float(e.max()), 6)
            row["energy_cv"]     = round(float(e.std() / e.mean()) if e.mean() > 0 else 0, 4)
        if "tempo_bpm" in acoustic_df.columns:
            valid = acoustic_df[acoustic_df["tempo_bpm"] > 0]["tempo_bpm"]
            row["tempo_mean"]    = round(float(valid.mean()), 2) if not valid.empty else None
            row["tempo_std"]     = round(float(valid.std()), 2) if not valid.empty else None
        if "pitch_mean" in acoustic_df.columns:
            valid = acoustic_df[acoustic_df["pitch_mean"] > 0]["pitch_mean"]
            row["pitch_mean"]    = round(float(valid.mean()), 2) if not valid.empty else None
    else:
        for k in ["energy_mean","energy_std","energy_max","energy_cv","tempo_mean","tempo_std","pitch_mean"]:
            row[k] = None

    # ── NLP features ─────────────────────────────────────────────────────────
    if transcript:
        from collections import Counter
        import re

        words = transcript.split()
        stop_words = {
            "the","a","an","and","or","but","in","on","at","to","for","of","with","is",
            "it","this","that","was","are","be","have","i","you","he","she","we","they",
            "do","not","so","if","as","my","your","his","her","um","uh","yeah","just","like",
        }
        clean_words = [w.lower().strip(".,!?\"'") for w in words if len(w.strip(".,!?\"'")) > 2]
        content_words = [w for w in clean_words if w not in stop_words]

        total = len(words)
        unique = len(set(clean_words))
        sentences = [s.strip() for s in re.split(r'[.!?]', transcript) if s.strip()]

        row["total_words_raw"]      = total
        row["unique_words"]         = unique
        row["lexical_diversity"]    = round(unique / total, 4) if total > 0 else 0
        row["avg_sentence_length"]  = round(total / len(sentences), 2) if sentences else 0
        row["n_sentences"]          = len(sentences)

        # Topic signals
        topic_seeds = {
            "topic_conflict":     ["fight","argue","conflict","crisis","angry","tension","blame","threat","war","enemy"],
            "topic_emotion":      ["feel","love","fear","hope","sad","happy","hurt","worried","care","trust","emotion"],
            "topic_money":        ["money","deal","cost","pay","profit","market","budget","revenue","invest","price","financial"],
            "topic_technology":   ["technology","data","system","digital","app","software","algorithm","ai","machine","code","tech"],
            "topic_family":       ["family","mother","father","sister","brother","marriage","relationship","child","home","together"],
            "topic_power":        ["power","control","government","leader","vote","decision","authority","policy","law","force"],
            "topic_health":       ["health","medical","doctor","disease","treatment","hospital","mental","care","patient","medicine"],
            "topic_education":    ["school","learn","teach","student","university","research","study","knowledge","education","class"],
        }
        lower_tx = transcript.lower()
        for topic, kws in topic_seeds.items():
            row[topic] = sum(lower_tx.count(kw) for kw in kws)

        # Top 5 content words
        freq = Counter(content_words)
        top5 = [w for w, _ in freq.most_common(5)]
        row["top_words"] = ", ".join(top5)
    else:
        for k in ["total_words_raw","unique_words","lexical_diversity","avg_sentence_length","n_sentences",
                  "topic_conflict","topic_emotion","topic_money","topic_technology","topic_family",
                  "topic_power","topic_health","topic_education","top_words"]:
            row[k] = None

    return row


def bundles_to_dataframe(items: List[Dict]) -> pd.DataFrame:
    """Convert a list of {metadata, bundle} dicts to a feature DataFrame."""
    rows = []
    for item in items:
        try:
            row = bundle_to_feature_row(item.get("bundle", {}), item.get("metadata", {}))
            rows.append(row)
        except Exception as e:
            rows.append({"title": item.get("metadata", {}).get("title", "error"), "error": str(e)})
    return pd.DataFrame(rows)


# ════════════════════════════════════════════════════════════════════════════
# OLS REGRESSION
# ════════════════════════════════════════════════════════════════════════════

def run_ols(df: pd.DataFrame, outcome: str, predictors: List[str]) -> Dict[str, Any]:
    """OLS regression with heteroskedasticity-robust standard errors."""
    try:
        import statsmodels.api as sm
        clean = df[[outcome] + predictors].dropna()
        if len(clean) < len(predictors) + 3:
            return {"error": f"Insufficient observations ({len(clean)}) for {len(predictors)} predictors."}
        X = sm.add_constant(clean[predictors])
        y = clean[outcome]
        model = sm.OLS(y, X).fit(cov_type="HC3")
        return {
            "model": "OLS",
            "outcome": outcome,
            "n": int(model.nobs),
            "r2": round(float(model.rsquared), 4),
            "r2_adj": round(float(model.rsquared_adj), 4),
            "f_stat": round(float(model.fvalue), 4) if model.fvalue else None,
            "f_pvalue": round(float(model.f_pvalue), 4) if model.f_pvalue else None,
            "coefficients": {
                var: {
                    "coef": round(float(model.params[var]), 6),
                    "se":   round(float(model.bse[var]), 6),
                    "t":    round(float(model.tvalues[var]), 4),
                    "p":    round(float(model.pvalues[var]), 4),
                    "sig":  _sig_stars(float(model.pvalues[var])),
                }
                for var in model.params.index
            },
            "error": None,
        }
    except Exception as e:
        return {"error": str(e)}


# ════════════════════════════════════════════════════════════════════════════
# FIXED EFFECTS PANEL MODEL
# ════════════════════════════════════════════════════════════════════════════

def run_fixed_effects(
    df: pd.DataFrame,
    outcome: str,
    predictors: List[str],
    entity_col: str = "uploader",
    time_col: str = "year",
) -> Dict[str, Any]:
    """
    Within-estimator fixed effects using entity and time demeaning.
    Requires linearmodels if available, falls back to OLS with dummies.
    """
    try:
        try:
            from linearmodels.panel import PanelOLS
            import statsmodels.api as sm

            panel_df = df[[outcome, entity_col, time_col] + predictors].dropna().copy()
            if len(panel_df) < 10:
                return {"error": "Need at least 10 observations for panel model."}

            panel_df = panel_df.set_index([entity_col, time_col])
            X = sm.add_constant(panel_df[predictors])
            model = PanelOLS(panel_df[outcome], X, entity_effects=True, time_effects=True).fit(
                cov_type="clustered", cluster_entity=True
            )
            return {
                "model": "Fixed Effects Panel (Two-Way)",
                "outcome": outcome,
                "entity": entity_col,
                "time": time_col,
                "n": int(model.nobs),
                "r2_within": round(float(model.rsquared), 4),
                "coefficients": {
                    var: {
                        "coef": round(float(model.params[var]), 6),
                        "se":   round(float(model.std_errors[var]), 6),
                        "t":    round(float(model.tstats[var]), 4),
                        "p":    round(float(model.pvalues[var]), 4),
                        "sig":  _sig_stars(float(model.pvalues[var])),
                    }
                    for var in model.params.index if var != "const"
                },
                "error": None,
            }
        except ImportError:
            # Fallback: OLS with entity dummy variables
            import statsmodels.api as sm
            clean = df[[outcome, entity_col] + predictors].dropna().copy()
            dummies = pd.get_dummies(clean[entity_col], prefix="fe", drop_first=True)
            X = pd.concat([clean[predictors], dummies], axis=1)
            X = sm.add_constant(X)
            model = sm.OLS(clean[outcome], X).fit(cov_type="HC3")
            coefs = {
                var: {
                    "coef": round(float(model.params[var]), 6),
                    "p":    round(float(model.pvalues[var]), 4),
                    "sig":  _sig_stars(float(model.pvalues[var])),
                }
                for var in predictors if var in model.params.index
            }
            return {
                "model": "Fixed Effects (OLS + Entity Dummies)",
                "n": int(model.nobs),
                "r2": round(float(model.rsquared), 4),
                "coefficients": coefs,
                "error": None,
            }
    except Exception as e:
        return {"error": str(e)}


# ════════════════════════════════════════════════════════════════════════════
# DIFFERENCE-IN-DIFFERENCES
# ════════════════════════════════════════════════════════════════════════════

def run_did(
    df: pd.DataFrame,
    outcome: str,
    treatment_col: str,
    post_col: str,
    controls: Optional[List[str]] = None,
) -> Dict[str, Any]:
    """
    Standard 2x2 DiD: outcome ~ treated + post + treated*post + controls
    treatment_col: 0/1 indicating treatment group
    post_col: 0/1 indicating post-treatment period
    """
    try:
        import statsmodels.api as sm
        cols = [outcome, treatment_col, post_col] + (controls or [])
        clean = df[cols].dropna().copy()
        if len(clean) < 10:
            return {"error": "Need at least 10 observations for DiD."}

        clean["interaction"] = clean[treatment_col] * clean[post_col]
        regressors = [treatment_col, post_col, "interaction"] + (controls or [])
        X = sm.add_constant(clean[regressors])
        model = sm.OLS(clean[outcome], X).fit(cov_type="HC3")

        did_coef = model.params.get("interaction", None)
        did_p    = model.pvalues.get("interaction", None)

        return {
            "model": "Difference-in-Differences",
            "outcome": outcome,
            "n": int(model.nobs),
            "r2": round(float(model.rsquared), 4),
            "did_estimate": round(float(did_coef), 6) if did_coef is not None else None,
            "did_pvalue":   round(float(did_p), 4) if did_p is not None else None,
            "did_sig":      _sig_stars(float(did_p)) if did_p is not None else "",
            "interpretation": (
                f"The treatment effect (post-treatment change for treated vs control) is "
                f"{did_coef:.4f} (p={did_p:.3f}){_sig_stars(did_p)}."
            ) if did_coef is not None else "Could not estimate DiD.",
            "coefficients": {
                var: {
                    "coef": round(float(model.params[var]), 6),
                    "p":    round(float(model.pvalues[var]), 4),
                    "sig":  _sig_stars(float(model.pvalues[var])),
                }
                for var in model.params.index
            },
            "error": None,
        }
    except Exception as e:
        return {"error": str(e)}


# ════════════════════════════════════════════════════════════════════════════
# RANDOM FOREST / XGBOOST
# ════════════════════════════════════════════════════════════════════════════

def run_random_forest(
    df: pd.DataFrame,
    outcome: str,
    predictors: List[str],
    n_estimators: int = 200,
    task: str = "regression",
) -> Dict[str, Any]:
    """Random forest with feature importance. task: 'regression' or 'classification'."""
    try:
        from sklearn.ensemble import RandomForestRegressor, RandomForestClassifier
        from sklearn.model_selection import cross_val_score
        from sklearn.preprocessing import LabelEncoder

        clean = df[[outcome] + predictors].dropna().copy()
        if len(clean) < 10:
            return {"error": "Need at least 10 observations."}

        X = clean[predictors].values
        y_raw = clean[outcome].values

        if task == "classification":
            le = LabelEncoder()
            y = le.fit_transform(y_raw)
            model = RandomForestClassifier(n_estimators=n_estimators, random_state=42, n_jobs=-1)
            cv_scores = cross_val_score(model, X, y, cv=min(5, len(clean)//2), scoring="accuracy")
            metric_name = "cv_accuracy"
        else:
            y = y_raw.astype(float)
            model = RandomForestRegressor(n_estimators=n_estimators, random_state=42, n_jobs=-1)
            cv_scores = cross_val_score(model, X, y, cv=min(5, len(clean)//2), scoring="r2")
            metric_name = "cv_r2"

        model.fit(X, y)
        importances = pd.Series(model.feature_importances_, index=predictors).sort_values(ascending=False)

        return {
            "model": "Random Forest",
            "task": task,
            "outcome": outcome,
            "n": len(clean),
            "n_estimators": n_estimators,
            metric_name: round(float(cv_scores.mean()), 4),
            f"{metric_name}_std": round(float(cv_scores.std()), 4),
            "feature_importance": {k: round(float(v), 6) for k, v in importances.items()},
            "top_predictors": list(importances.head(5).index),
            "error": None,
        }
    except Exception as e:
        return {"error": str(e)}


def run_xgboost(
    df: pd.DataFrame,
    outcome: str,
    predictors: List[str],
    task: str = "regression",
) -> Dict[str, Any]:
    """XGBoost with early stopping and feature importance."""
    try:
        import xgboost as xgb
        from sklearn.model_selection import cross_val_score
        from sklearn.preprocessing import LabelEncoder

        clean = df[[outcome] + predictors].dropna().copy()
        if len(clean) < 10:
            return {"error": "Need at least 10 observations."}

        X = clean[predictors].values
        y_raw = clean[outcome].values

        if task == "classification":
            le = LabelEncoder()
            y = le.fit_transform(y_raw)
            model = xgb.XGBClassifier(n_estimators=200, learning_rate=0.05,
                                       max_depth=4, random_state=42,
                                       eval_metric="logloss", verbosity=0)
            cv_scores = cross_val_score(model, X, y, cv=min(5, len(clean)//2), scoring="accuracy")
            metric_name = "cv_accuracy"
        else:
            y = y_raw.astype(float)
            model = xgb.XGBRegressor(n_estimators=200, learning_rate=0.05,
                                      max_depth=4, random_state=42,
                                      eval_metric="rmse", verbosity=0)
            cv_scores = cross_val_score(model, X, y, cv=min(5, len(clean)//2), scoring="r2")
            metric_name = "cv_r2"

        model.fit(X, y)
        importances = pd.Series(model.feature_importances_, index=predictors).sort_values(ascending=False)

        return {
            "model": "XGBoost",
            "task": task,
            "outcome": outcome,
            "n": len(clean),
            metric_name: round(float(cv_scores.mean()), 4),
            f"{metric_name}_std": round(float(cv_scores.std()), 4),
            "feature_importance": {k: round(float(v), 6) for k, v in importances.items()},
            "top_predictors": list(importances.head(5).index),
            "error": None,
        }
    except ImportError:
        return {"error": "xgboost not installed. Run: pip install xgboost"}
    except Exception as e:
        return {"error": str(e)}


# ════════════════════════════════════════════════════════════════════════════
# LDA TOPIC MODELING
# ════════════════════════════════════════════════════════════════════════════

def run_lda(
    transcripts: List[str],
    titles: List[str],
    n_topics: int = 5,
    n_top_words: int = 10,
) -> Dict[str, Any]:
    """LDA topic modeling on a corpus of transcripts."""
    try:
        from sklearn.feature_extraction.text import CountVectorizer
        from sklearn.decomposition import LatentDirichletAllocation

        stop_words = "english"
        vectorizer = CountVectorizer(
            max_df=0.9, min_df=2, stop_words=stop_words,
            max_features=5000, ngram_range=(1, 2),
        )
        dtm = vectorizer.fit_transform(transcripts)
        vocab = vectorizer.get_feature_names_out()

        lda = LatentDirichletAllocation(
            n_components=n_topics, random_state=42,
            max_iter=20, learning_method="online",
        )
        doc_topics = lda.fit_transform(dtm)

        topics = {}
        for i, comp in enumerate(lda.components_):
            top_idxs = comp.argsort()[-n_top_words:][::-1]
            topics[f"Topic {i+1}"] = [vocab[j] for j in top_idxs]

        # Dominant topic per document
        doc_topic_df = pd.DataFrame(
            doc_topics,
            columns=[f"topic_{i+1}_weight" for i in range(n_topics)],
        )
        doc_topic_df["title"] = titles[:len(doc_topic_df)]
        doc_topic_df["dominant_topic"] = doc_topics.argmax(axis=1) + 1

        return {
            "model": "LDA Topic Model",
            "n_topics": n_topics,
            "n_documents": len(transcripts),
            "topics": topics,
            "doc_topic_weights": doc_topic_df.to_dict(orient="records"),
            "perplexity": round(float(lda.perplexity(dtm)), 2),
            "error": None,
        }
    except Exception as e:
        return {"error": str(e)}


# ════════════════════════════════════════════════════════════════════════════
# TIME SERIES TREND ANALYSIS
# ════════════════════════════════════════════════════════════════════════════

def run_time_series(
    df: pd.DataFrame,
    time_col: str,
    value_cols: List[str],
    group_col: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Trend analysis: linear trend fit, rolling averages, changepoint detection.
    """
    try:
        import statsmodels.api as sm

        results = {}

        work_df = df[[time_col] + value_cols + ([group_col] if group_col else [])].dropna(subset=[time_col]).copy()
        work_df = work_df.sort_values(time_col)

        # Convert time to numeric if needed
        if work_df[time_col].dtype == object:
            try:
                work_df[time_col] = pd.to_datetime(work_df[time_col])
                work_df["_t"] = (work_df[time_col] - work_df[time_col].min()).dt.days
            except Exception:
                work_df["_t"] = range(len(work_df))
        else:
            work_df["_t"] = work_df[time_col]

        for col in value_cols:
            sub = work_df[["_t", col]].dropna()
            if len(sub) < 3:
                continue

            X = sm.add_constant(sub["_t"])
            model = sm.OLS(sub[col], X).fit()
            slope = float(model.params["_t"])
            p = float(model.pvalues["_t"])

            # Rolling 3-period average
            rolling = sub[col].rolling(window=min(3, len(sub)), min_periods=1).mean().tolist()

            results[col] = {
                "n": len(sub),
                "trend_slope": round(slope, 6),
                "trend_pvalue": round(p, 4),
                "trend_sig": _sig_stars(p),
                "trend_direction": "↑ Increasing" if slope > 0 else "↓ Decreasing",
                "r2": round(float(model.rsquared), 4),
                "rolling_avg": [round(v, 4) for v in rolling],
                "mean": round(float(sub[col].mean()), 4),
                "std": round(float(sub[col].std()), 4),
            }

        return {"model": "Time Series Trend", "results": results, "error": None}
    except Exception as e:
        return {"error": str(e)}


# ════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ════════════════════════════════════════════════════════════════════════════

def export_to_excel(
    feature_df: pd.DataFrame,
    model_results: Optional[Dict] = None,
    lda_results: Optional[Dict] = None,
    output_path: str = "academic_analysis.xlsx",
) -> str:
    """
    Export full academic analysis to a multi-sheet Excel workbook.
    Sheet 1: Feature Dataset
    Sheet 2: Descriptive Statistics
    Sheet 3: Model Results
    Sheet 4: LDA Topics
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Colors ───────────────────────────────────────────────────────────────
    HEADER_FILL = PatternFill("solid", start_color="1E90FF", end_color="1E90FF")
    SUBHEADER_FILL = PatternFill("solid", start_color="D3E8FF", end_color="D3E8FF")
    ALT_FILL = PatternFill("solid", start_color="F5F9FF", end_color="F5F9FF")
    WHITE_FILL = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    BOLD_FONT = Font(bold=True, name="Arial", size=10)
    NORMAL_FONT = Font(name="Arial", size=10)
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(cell, fill=HEADER_FILL, font=HEADER_FONT):
        cell.fill = fill
        cell.font = font
        cell.alignment = CENTER
        cell.border = BORDER

    def style_cell(cell, font=NORMAL_FONT, fill=WHITE_FILL, align=LEFT):
        cell.font = font
        cell.fill = fill
        cell.alignment = align
        cell.border = BORDER

    def autofit(ws, min_w=10, max_w=50):
        for col in ws.columns:
            length = max((len(str(cell.value or "")) for cell in col), default=min_w)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 1: Feature Dataset
    # ════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Feature Dataset"

    for col_idx, col_name in enumerate(feature_df.columns, 1):
        cell = ws1.cell(row=1, column=col_idx, value=col_name)
        style_header(cell)

    for row_idx, row_data in enumerate(feature_df.itertuples(index=False), 2):
        fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL
        for col_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            style_cell(cell, fill=fill)

    ws1.freeze_panes = "A2"
    autofit(ws1)

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 2: Descriptive Statistics
    # ════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Descriptive Statistics")
    numeric_cols = feature_df.select_dtypes(include=[np.number]).columns.tolist()
    if numeric_cols:
        desc = feature_df[numeric_cols].describe().round(4)
        desc_reset = desc.reset_index()
        for col_idx, col_name in enumerate(desc_reset.columns, 1):
            cell = ws2.cell(row=1, column=col_idx, value=col_name)
            style_header(cell)
        for row_idx, row_data in enumerate(desc_reset.itertuples(index=False), 2):
            fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL
            for col_idx, val in enumerate(row_data, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=val)
                style_cell(cell, fill=fill)
        ws2.freeze_panes = "B2"
        autofit(ws2)

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 3: Model Results
    # ════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Model Results")
    r = 1

    if model_results:
        for model_name, result in model_results.items():
            if result.get("error"):
                ws3.cell(row=r, column=1, value=f"{model_name}: ERROR — {result['error']}").font = Font(color="FF0000", name="Arial")
                r += 2
                continue

            # Model header
            cell = ws3.cell(row=r, column=1, value=model_name)
            style_header(cell)
            ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            r += 1

            # Model-level stats
            for key in ["model","outcome","n","r2","r2_adj","r2_within","did_estimate","did_pvalue","did_sig","perplexity"]:
                if key in result and result[key] is not None:
                    ws3.cell(row=r, column=1, value=key).font = BOLD_FONT
                    ws3.cell(row=r, column=2, value=result[key]).font = NORMAL_FONT
                    r += 1

            # Coefficients table
            coefs = result.get("coefficients") or result.get("feature_importance")
            if coefs:
                headers = ["Variable", "Coefficient / Importance", "Std Error", "t-stat", "p-value", "Significance"]
                for col_idx, h in enumerate(headers, 1):
                    cell = ws3.cell(row=r, column=col_idx, value=h)
                    style_header(cell, fill=SUBHEADER_FILL, font=BOLD_FONT)
                r += 1
                for var, stats in coefs.items():
                    if isinstance(stats, dict):
                        vals = [var, stats.get("coef"), stats.get("se"), stats.get("t"), stats.get("p"), stats.get("sig","")]
                    else:
                        vals = [var, round(float(stats), 6), "", "", "", ""]
                    fill = ALT_FILL if r % 2 == 0 else WHITE_FILL
                    for col_idx, val in enumerate(vals, 1):
                        style_cell(ws3.cell(row=r, column=col_idx, value=val), fill=fill)
                    r += 1
            r += 2

    ws3.freeze_panes = "A2"
    autofit(ws3)

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 4: LDA Topics
    # ════════════════════════════════════════════════════════════════════════
    if lda_results and not lda_results.get("error"):
        ws4 = wb.create_sheet("LDA Topics")
        r = 1

        cell = ws4.cell(row=r, column=1, value="LDA Topic Model Results")
        style_header(cell)
        ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r += 1

        for key in ["n_topics", "n_documents", "perplexity"]:
            if key in lda_results:
                ws4.cell(row=r, column=1, value=key).font = BOLD_FONT
                ws4.cell(row=r, column=2, value=lda_results[key]).font = NORMAL_FONT
                r += 1
        r += 1

        # Topic keywords
        cell = ws4.cell(row=r, column=1, value="Topic Keywords")
        style_header(cell)
        ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r += 1
        for topic, words in lda_results.get("topics", {}).items():
            ws4.cell(row=r, column=1, value=topic).font = BOLD_FONT
            ws4.cell(row=r, column=2, value=", ".join(words)).font = NORMAL_FONT
            r += 1
        r += 1

        # Document-topic weights
        if lda_results.get("doc_topic_weights"):
            cell = ws4.cell(row=r, column=1, value="Document Topic Weights")
            style_header(cell)
            ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
            r += 1
            dtw_df = pd.DataFrame(lda_results["doc_topic_weights"])
            for col_idx, col_name in enumerate(dtw_df.columns, 1):
                style_header(ws4.cell(row=r, column=col_idx, value=col_name), fill=SUBHEADER_FILL, font=BOLD_FONT)
            r += 1
            for row_idx, row_data in enumerate(dtw_df.itertuples(index=False), 0):
                fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL
                for col_idx, val in enumerate(row_data, 1):
                    if isinstance(val, float):
                        val = round(val, 4)
                    style_cell(ws4.cell(row=r, column=col_idx, value=val), fill=fill)
                r += 1

        ws4.freeze_panes = "A2"
        autofit(ws4)

    wb.save(output_path)
    return output_path


# ════════════════════════════════════════════════════════════════════════════
# HELPERS
# ════════════════════════════════════════════════════════════════════════════

def _sig_stars(p: float) -> str:
    if p < 0.001: return "***"
    if p < 0.01:  return "**"
    if p < 0.05:  return "*"
    if p < 0.10:  return "."
    return ""


def get_numeric_cols(df: pd.DataFrame) -> List[str]:
    return df.select_dtypes(include=[np.number]).columns.tolist()
